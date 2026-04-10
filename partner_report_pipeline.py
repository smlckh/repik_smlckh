"""
Partner Contract Report — Filter & SharePoint Upload
======================================================
Reads the latest Excel file from a local folder, keeps only the
configured columns, applies contract filtering logic, merges with
existing partner files (preserving notes), and uploads to SharePoint.

REQUIREMENTS:
    pip install pandas openpyxl office365-rest-python-client python-dotenv python-dateutil

SETUP:
    1. Fill in config.env
    2. Run manually via run_pipeline.bat, or schedule via Windows Task Scheduler
"""

import os
import json
import hashlib
import logging
import tempfile
import glob
from datetime import datetime
from dateutil.relativedelta import relativedelta
from dotenv import load_dotenv

import pandas as pd
from office365.sharepoint.client_context import ClientContext

# ─────────────────────────────────────────────
# LOAD CONFIG
# ─────────────────────────────────────────────

load_dotenv("config.env")

EXCEL_FOLDER     = os.getenv("EXCEL_FOLDER", r"C:\path\to\daily\excel\files")
SHAREPOINT_URL   = os.getenv("SHAREPOINT_URL", "https://cissrocz.sharepoint.com/sites/PE-Obchod")
SP_CLIENT_ID     = os.getenv("SP_CLIENT_ID", "")
SP_CLIENT_SECRET = os.getenv("SP_CLIENT_SECRET", "")
SP_TENANT_ID     = os.getenv("SP_TENANT_ID", "")
BASE_FOLDER      = os.getenv("BASE_FOLDER", "Sdilene dokumenty/Partner Reports")
MASTER_FOLDER    = os.getenv("MASTER_FOLDER", "Sdilene dokumenty/Master")
MASTER_FILENAME  = "master.xlsx"
HASH_FILE        = os.getenv("HASH_FILE", "partner_hashes.json")
LOG_FILE         = os.getenv("LOG_FILE", "upload_log.txt")

OUTPUT_FILENAME  = "report_presmluvneni.xlsx"

# ─────────────────────────────────────────────
# COLUMNS TO KEEP FROM DAILY EXCEL
# ─────────────────────────────────────────────

COLUMNS_TO_KEEP = [
    "Obor kód",
    "Kategorie segment",
    "Č. zákazníka",
    "Jméno",
    "Příjmení",
    "Název firmy",
    "IČ",
    "Datum narození",
    "Mobil",
    "Telefon",
    "E-mailová adresa",
    "Město_1", "PSČ_1", "Ulice_1", "Č. popisné_1", "Č. orientační_1",
    "Způsob obsluhy",
    "Nadřazený obchodník",
    "Č. smluvního účtu",
    "Číslo smlouvy",
    "Stav smlouvy",
    "Důvod stavu smlouvy",
    "Sml. platí od",
    "Sml. platí do",
    "Platnost posledního dodatku do",
    "Souhlas GDPR",
    "Prolongace",
    "Lhůta prolongace",
    "Produkt název",
    "EAN/EIC odb. místa",
    "Předpokládaná roční spotřeba",
    "Roční spotřeba OTE - VT / Plyn (kWh)",
    "Roční spotřeba OTE - NT (kWh)",
    "Komoditní sazba - VT/ZP (MWh)",
    "Komoditní sazba - NT (MWh)",
    "Měna MWh",
    "Přirážka",
    "Měna přirážka",
    "Kapacitní sazba (měsíc)",
    "Měna kapacita",
    "Č. místa",
    "Sdružená fakturace",
    "Fakturační cyklus",
    "Datum založení smlouvy",
    "Datum podpisu smlouvy",
]

# Script-managed columns
COL_STATUS        = "Status"
COL_YOUR_NOTES    = "Your Notes"
COL_PARTNER_NOTES = "Partner Notes"

# Key columns
COL_PARTNER     = "Způsob obsluhy"
COL_END_DATE    = "Platnost posledního dodatku do"
COL_PRODUCT     = "Produkt název"
COL_CONTRACT_ID = "Číslo smlouvy"

# ─────────────────────────────────────────────
# PREBIRAME — partners who handle own renewals
# ─────────────────────────────────────────────

PREBIRAME = [
    "7500", "7020", "4000", "5501", "7049", "9700", "7051", "7035",
    "5506", "9900", "6080", "7032", "90090-RE Veřejná aukční",
    "7043", "7031", "7010", "3600", "7053", "7036", "7039", "7037",
    "7041", "7033", "7050", "7040", "4702"
]

# Always Ours — never partner contracts, no SharePoint folder
ALWAYS_OURS = [
    "9000 ZPRACOVÁNO BO",
    "90016 BO Smlouvy YD",
]

# ─────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(LOG_FILE, encoding="utf-8")
    ]
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# EXCEL FORMATTING
# ─────────────────────────────────────────────

def format_as_excel_table(df: pd.DataFrame, tmp_path: str, table_name: str = "Table1"):
    """Saves dataframe as a formatted Excel table with frozen header and auto widths."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    df.to_excel(tmp_path, index=False)
    wb = load_workbook(tmp_path)
    ws = wb.active

    # Create table
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(tab)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    wb.save(tmp_path)


# ─────────────────────────────────────────────
# LOAD & CLEAN DAILY EXCEL
# ─────────────────────────────────────────────

def load_latest_excel() -> pd.DataFrame:
    """Picks the most recently modified .xlsx from EXCEL_FOLDER."""
    files = glob.glob(os.path.join(EXCEL_FOLDER, "*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No .xlsx files found in: {EXCEL_FOLDER}")
    latest = max(files, key=os.path.getmtime)
    log.info(f"Loading: {latest}")

    # header=1 because column names are in row 2
    df = pd.read_excel(latest, header=1, engine="openpyxl")

    # Rename duplicate address columns → Město_1, Město_2, Město_3 etc.
    dup_cols = ["Město", "PSČ", "Ulice", "Č. popisné", "Č. orientační"]
    counters = {}
    new_cols = []
    for col in list(df.columns):
        if col in dup_cols:
            counters[col] = counters.get(col, 0) + 1
            new_cols.append(f"{col}_{counters[col]}")
        else:
            new_cols.append(col)
    df.columns = new_cols

    # Keep only configured columns
    available = [c for c in COLUMNS_TO_KEEP if c in df.columns]
    missing   = [c for c in COLUMNS_TO_KEEP if c not in df.columns]
    if missing:
        log.warning(f"Columns not found in source Excel (skipped): {missing}")

    df = df[available].copy()

    # Č. zákazníka as string (avoid 12345.0 formatting)
    if "Č. zákazníka" in df.columns:
        df["Č. zákazníka"] = df["Č. zákazníka"].astype(str).str.replace(r'\.0$', '', regex=True)

    df[COL_END_DATE] = pd.to_datetime(df[COL_END_DATE], errors="coerce")

    # Keep only active contracts
    before = len(df)
    df = df[df["Stav smlouvy"] == "Účinná"].reset_index(drop=True)
    log.info(f"Filtered to 'Účinná': {len(df)} rows (dropped {before - len(df)})")

    # Exclude contracts ending beyond 12 months
    today   = pd.Timestamp.today().normalize()
    enddate = today + relativedelta(months=12)
    before  = len(df)
    df = df[df[COL_END_DATE] <= enddate].reset_index(drop=True)
    log.info(f"Excluded >12 months: {len(df)} rows (dropped {before - len(df)})")

    log.info(f"Loaded {len(df)} rows, {len(available)} columns.")
    return df


# ─────────────────────────────────────────────
# STATUS CALCULATION (vectorized)
# ─────────────────────────────────────────────

def calculate_status_vectorized(df: pd.DataFrame) -> pd.Series:
    """
    Based on DAX measure 'k presmluvneni PCE' where flag 1 = Ours, flag 0 = Partner.

    Partner:           PREBIRAME, ends 3-12 months, not ALWAYS_OURS, not SPOT/Výrobna
    Transferred to us: PREBIRAME, ends <3 months
    Ours:              everything else
    """
    today     = pd.Timestamp.today().normalize()
    startdate = today + relativedelta(months=3)
    enddate   = today + relativedelta(months=12)

    doddo   = pd.to_datetime(df[COL_END_DATE], errors="coerce")
    partner = df[COL_PARTNER].fillna("").astype(str)
    product = df[COL_PRODUCT].fillna("").astype(str)

    always_ours_mask = pd.Series(False, index=df.index)
    for code in ALWAYS_OURS:
        always_ours_mask |= partner.str.contains(code, regex=False)

    prebirame_mask = pd.Series(False, index=df.index)
    for code in PREBIRAME:
        prebirame_mask |= partner.str.contains(code, regex=False)

    is_spot   = product.str.contains("SPOT", regex=False) | \
                product.str.contains("Výrobna", regex=False)
    in_window = (doddo >= startdate) & (doddo < enddate)
    under_3   = doddo < startdate

    status = pd.Series("Ours", index=df.index)
    status[prebirame_mask & in_window & ~always_ours_mask & ~is_spot] = "Partner"
    status[prebirame_mask & under_3   & ~always_ours_mask & ~is_spot] = "Transferred to us"

    return status


# ─────────────────────────────────────────────
# SHAREPOINT HELPERS
# ─────────────────────────────────────────────

def get_sharepoint_context() -> ClientContext:
    from office365.runtime.auth.client_credential import ClientCredential
    credentials = ClientCredential(SP_CLIENT_ID, SP_CLIENT_SECRET)
    return ClientContext(SHAREPOINT_URL).with_credentials(credentials)


def site_relative_url(folder: str) -> str:
    from urllib.parse import urlparse
    site_path = urlparse(SHAREPOINT_URL).path.rstrip("/")
    return f"{site_path}/{folder}"


def upload_file_to_sharepoint(ctx: ClientContext, folder: str, filename: str, file_bytes: bytes):
    """Uploads a file to a SharePoint folder, creating subfolders if needed."""
    from urllib.parse import urlparse
    site_path = urlparse(SHAREPOINT_URL).path.rstrip("/")

    parts    = folder.split("/")
    base     = "/".join(parts[:2])
    sub      = "/".join(parts[2:])
    base_srv = f"{site_path}/{base}"

    if sub:
        base_folder = ctx.web.get_folder_by_server_relative_url(base_srv)
        base_folder.folders.add(sub).execute_query()
        target_srv = f"{base_srv}/{sub}"
    else:
        target_srv = base_srv

    ctx.web.get_folder_by_server_relative_url(target_srv) \
        .upload_file(filename, file_bytes).execute_query()


def download_file_from_sharepoint(ctx: ClientContext, folder: str, filename: str):
    """Downloads a file from SharePoint to a temp path. Returns path or None."""
    try:
        srv_url = site_relative_url(f"{folder}/{filename}")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        with open(tmp_path, "wb") as f:
            ctx.web.get_file_by_server_relative_url(srv_url).download(f).execute_query()
        return tmp_path
    except Exception as e:
        log.debug(f"  Could not download {folder}/{filename}: {e}")
        return None


def safe_folder_name(name: str) -> str:
    for ch in r'\/:*?"<>|#%':
        name = name.replace(ch, "_")
    return name.strip()


def compute_hash(df: pd.DataFrame) -> str:
    return hashlib.md5(pd.util.hash_pandas_object(df, index=True).values).hexdigest()


def load_hashes() -> dict:
    if os.path.exists(HASH_FILE):
        with open(HASH_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_hashes(hashes: dict):
    with open(HASH_FILE, "w", encoding="utf-8") as f:
        json.dump(hashes, f, ensure_ascii=False, indent=2)


def sync_partner_notes_to_master(ctx: ClientContext, master_df: pd.DataFrame) -> pd.DataFrame:
    """
    Reads partner notes from all partner files on SharePoint and syncs into master.
    Lists actual folders from SharePoint instead of guessing folder names.
    """
    log.info("Syncing partner notes back to master...")
    from urllib.parse import urlparse
    site_path = urlparse(SHAREPOINT_URL).path.rstrip("/")
    base_srv  = f"{site_path}/{BASE_FOLDER}"

    try:
        folders = ctx.web.get_folder_by_server_relative_url(base_srv) \
                      .folders.get().execute_query()
    except Exception as e:
        log.error(f"Could not list partner folders: {e}")
        return master_df

    total_notes = 0
    for folder in folders:
        file_url = f"{folder.serverRelativeUrl}/{OUTPUT_FILENAME}"
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp_path = tmp.name
            with open(tmp_path, "wb") as f:
                ctx.web.get_file_by_server_relative_url(file_url).download(f).execute_query()

            existing = pd.read_excel(tmp_path)
            os.unlink(tmp_path)

            if COL_CONTRACT_ID in existing.columns and COL_PARTNER_NOTES in existing.columns:
                notes = existing.set_index(COL_CONTRACT_ID)[COL_PARTNER_NOTES].dropna().to_dict()
                notes = {k: v for k, v in notes.items() if str(v).strip()}
                for contract_id, note in notes.items():
                    master_df.loc[master_df[COL_CONTRACT_ID] == contract_id, COL_PARTNER_NOTES] = note
                if notes:
                    log.info(f"  Synced {len(notes)} notes from {folder.name}")
                    total_notes += len(notes)
        except Exception as e:
            log.debug(f"  Could not read {file_url}: {e}")

    log.info(f"Total notes synced: {total_notes}")

    # Re-upload master with updated notes
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        format_as_excel_table(master_df, tmp_path, "Master")
        with open(tmp_path, "rb") as f:
            file_bytes = f.read()
        upload_file_to_sharepoint(ctx, MASTER_FOLDER, MASTER_FILENAME, file_bytes)
        log.info("Master updated with latest partner notes.")
    finally:
        os.unlink(tmp_path)

    return master_df


# ─────────────────────────────────────────────
# MASTER FILE
# ─────────────────────────────────────────────

def update_master(new_df: pd.DataFrame, ctx: ClientContext) -> pd.DataFrame:
    """
    Downloads master from SharePoint, merges new data (preserving Your Notes),
    recalculates Status, uploads back to SharePoint.
    """
    tmp_path = download_file_from_sharepoint(ctx, MASTER_FOLDER, MASTER_FILENAME)

    if tmp_path:
        master = pd.read_excel(tmp_path)
        os.unlink(tmp_path)
        master[COL_END_DATE] = pd.to_datetime(master[COL_END_DATE], errors="coerce")
        if COL_YOUR_NOTES not in master.columns:
            master[COL_YOUR_NOTES] = ""
        log.info(f"Existing master loaded from SharePoint: {len(master)} rows.")

        notes  = master[[COL_CONTRACT_ID, COL_YOUR_NOTES]].copy()
        merged = new_df.merge(notes, on=COL_CONTRACT_ID, how="left")
        merged[COL_YOUR_NOTES] = merged[COL_YOUR_NOTES].fillna("")

        before  = len(merged)
        merged  = merged[merged["Stav smlouvy"] == "Účinná"].reset_index(drop=True)
        dropped = before - len(merged)
        if dropped:
            log.info(f"Removed {dropped} contracts no longer Účinná from master.")
    else:
        log.info("No master file found on SharePoint — creating fresh.")
        merged = new_df.copy()
        merged[COL_YOUR_NOTES] = ""

    merged[COL_STATUS] = calculate_status_vectorized(merged)

    # Ensure Partner Notes column exists
    if COL_PARTNER_NOTES not in merged.columns:
        merged[COL_PARTNER_NOTES] = ""

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        format_as_excel_table(merged, tmp_path, "Master")
        with open(tmp_path, "rb") as f:
            file_bytes = f.read()
        upload_file_to_sharepoint(ctx, MASTER_FOLDER, MASTER_FILENAME, file_bytes)
        log.info(f"Master saved to SharePoint: {len(merged)} rows → {MASTER_FOLDER}/{MASTER_FILENAME}")
    finally:
        os.unlink(tmp_path)

    return merged


# ─────────────────────────────────────────────
# PARTNER FILE UPLOAD
# ─────────────────────────────────────────────

def fetch_existing_partner_notes(ctx: ClientContext, folder: str) -> dict:
    """Downloads existing partner file and extracts {contract_id: note}."""
    notes = {}
    try:
        tmp_path = download_file_from_sharepoint(ctx, folder, OUTPUT_FILENAME)
        if not tmp_path:
            log.info("  No existing partner file — starting fresh.")
            return notes
        existing = pd.read_excel(tmp_path)
        os.unlink(tmp_path)
        if COL_CONTRACT_ID in existing.columns and COL_PARTNER_NOTES in existing.columns:
            notes = existing.set_index(COL_CONTRACT_ID)[COL_PARTNER_NOTES].dropna().to_dict()
            notes = {k: v for k, v in notes.items() if str(v).strip()}
        log.info(f"  Preserved {len(notes)} partner notes.")
    except Exception:
        log.info("  No existing partner file — starting fresh.")
    return notes


def upload_partner_file(
    ctx: ClientContext,
    partner: str,
    df: pd.DataFrame,
    previous_hashes: dict,
    new_hashes: dict
):
    partner_str  = str(partner)
    folder_name  = safe_folder_name(partner_str)
    folder       = f"{BASE_FOLDER}/{folder_name}"
    current_hash = compute_hash(df)
    new_hashes[partner_str] = current_hash

    if previous_hashes.get(partner_str) == current_hash:
        log.info("  No changes — skipping.")
        return

    existing_notes = fetch_existing_partner_notes(ctx, folder)

    partner_cols = [c for c in COLUMNS_TO_KEEP if c in df.columns] + [COL_STATUS]
    partner_df   = df[partner_cols].copy()
    partner_df[COL_PARTNER_NOTES] = partner_df[COL_CONTRACT_ID].map(existing_notes).fillna("")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        format_as_excel_table(partner_df, tmp_path, "Presmluvneni")
        with open(tmp_path, "rb") as f:
            file_bytes = f.read()
        upload_file_to_sharepoint(ctx, folder, OUTPUT_FILENAME, file_bytes)
        log.info(f"  ✓ {len(partner_df)} rows uploaded to {folder}/")
    finally:
        os.unlink(tmp_path)


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    log.info("=" * 60)
    log.info(f"Run started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # 1. Load latest daily Excel
    try:
        raw_df = load_latest_excel()
    except Exception as e:
        log.error(f"Failed to load Excel: {e}")
        return

    # 2. Connect to SharePoint
    log.info("Connecting to SharePoint...")
    try:
        ctx = get_sharepoint_context()
        ctx.web.get().execute_query()
        log.info("Connected.")
    except Exception as e:
        log.error(f"SharePoint connection failed: {e}")
        return

    # 3. Merge into master on SharePoint
    master_df = update_master(raw_df, ctx)

    # 4. Filter to Partner rows only
    partner_visible = master_df[master_df[COL_STATUS] == "Partner"].copy()
    log.info(f"Rows for partners (Status=Partner): {len(partner_visible)} / {len(master_df)} total")

    if partner_visible.empty:
        log.warning("No rows for partners after filtering. Nothing to upload.")
        return

    # 5. Upload per partner
    partners        = partner_visible[COL_PARTNER].dropna().unique()
    previous_hashes = load_hashes()
    new_hashes      = {}

    log.info(f"Processing {len(partners)} partners...")

    for partner in partners:
        partner_df = partner_visible[
            partner_visible[COL_PARTNER] == partner
        ].reset_index(drop=True)
        log.info(f"\nPartner: {partner} ({len(partner_df)} rows)")
        try:
            upload_partner_file(ctx, str(partner), partner_df, previous_hashes, new_hashes)
        except Exception as e:
            log.error(f"  Failed: {e}")
            new_hashes[str(partner)] = previous_hashes.get(str(partner), "")

    save_hashes(new_hashes)

    # 6. Sync partner notes back into master
    sync_partner_notes_to_master(ctx, master_df)

    log.info(f"\nAll done: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info("=" * 60)


if __name__ == "__main__":
    main()